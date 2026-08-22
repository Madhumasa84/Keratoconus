"""Local exports and guarded screen-positive referral PDF generation."""
from __future__ import annotations

import hashlib
import html
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .protocol import load_protocol

log = logging.getLogger(__name__)

DISCLAIMER = (
    "KeraScan is an experimental initial screening aid. This report does not diagnose or "
    "exclude keratoconus. A screen-positive result indicates that further ophthalmic "
    "evaluation is recommended outside this application."
)

MEASUREMENT_REASON_DETAILS = {
    "K2_ABOVE_46_8_D": ("K2", "k2_d", "Above study threshold"),
    "PACHYMETRY_BELOW_480_UM": ("Pachymetry", "pachymetry_um", "Below study threshold"),
    "CYLINDER_MAGNITUDE_ABOVE_1_5_D": ("Cylinder", "cylinder_d", "Above study threshold"),
}


def _safe(value: Any, default: str = "—") -> str:
    if value is None or value == "":
        return default
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _paragraph_text(value: Any, default: str = "—") -> str:
    """Escape all values that will enter ReportLab's XML-like Paragraph syntax."""
    return html.escape(_safe(value, default), quote=True)


def _sha256(path: str | Path) -> str | None:
    try:
        digest = hashlib.sha256()
        with Path(path).open("rb") as handle:
            for chunk in iter(lambda: handle.read(65536), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError:
        return None


class ReportService:
    """Generate local JSON/XLSX and a detailed PDF for a true REFER action only."""

    @staticmethod
    def _is_refer_action(screening_data: dict[str, Any]) -> bool:
        # A label alone is not sufficient. The persisted final child-level
        # action must explicitly be REFER before a detailed referral report may
        # exist.
        return screening_data.get("overall_action") == "REFER"

    @staticmethod
    def _eyes_by_laterality(screening_data: dict[str, Any]) -> dict[str, dict[str, Any]]:
        return {str(eye.get("laterality")): eye for eye in screening_data.get("eyes", []) if eye.get("laterality") in {"OD", "OS"}}

    @staticmethod
    def _latest_measurements(eye: dict[str, Any]) -> dict[str, Any]:
        measurements = eye.get("measurements") or []
        if not measurements:
            return {}
        return sorted(measurements, key=lambda item: (item.get("reading_number") or 0, item.get("created_at") or ""))[-1]

    @staticmethod
    def _eye_decision(eye: dict[str, Any]) -> str:
        decisions = eye.get("decisions") or []
        if decisions:
            return str(decisions[-1].get("final_result") or decisions[-1].get("automated_result") or "—")
        return str(eye.get("per_eye_decision") or "—")

    def _affected_eyes(self, screening_data: dict[str, Any]) -> list[str]:
        explicit = screening_data.get("affected_eyes")
        if isinstance(explicit, list):
            return [eye for eye in explicit if eye in {"OD", "OS"}]
        referral_decisions = {
            "HIGH_RISK_SCREEN_POSITIVE",
            "SCREEN_POSITIVE_IMAGE_ONLY",
            "DISCORDANT_SCREEN_POSITIVE",
        }
        return [
            eye.get("laterality") for eye in screening_data.get("eyes", [])
            if self._eye_decision(eye) in referral_decisions and eye.get("laterality") in {"OD", "OS"}
        ]

    # ------------------------------------------------------------------
    # JSON / workbook exports
    # ------------------------------------------------------------------

    def generate_json(self, screening_data: dict[str, Any], output_path: str) -> str:
        from app.services.privacy import redact_paths

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        export = {
            "export_generated_at": datetime.now(timezone.utc).isoformat(),
            "disclaimer": DISCLAIMER,
            "screening": redact_paths(screening_data),
        }
        output.write_text(json.dumps(export, indent=2, default=str), encoding="utf-8")
        return str(output.resolve())

    def generate_excel(self, screening_data: dict[str, Any], output_path: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:  # pragma: no cover - dependency is declared
            raise RuntimeError("openpyxl is required for local XLSX export.") from exc

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = openpyxl.Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["KeraScan Screening Record"])
        summary["A1"].font = Font(bold=True, size=14)
        for label, value in (
            ("Anonymous child/study ID", screening_data.get("screening_id")),
            ("Screening date", screening_data.get("screening_date")),
            ("Operator ID", screening_data.get("operator_id")),
            ("Device ID", screening_data.get("device_id")),
            ("Protocol version", screening_data.get("protocol_version")),
            ("Child result", screening_data.get("overall_result")),
            ("Child action", screening_data.get("overall_action")),
            ("Referral priority", screening_data.get("referral_priority")),
        ):
            summary.append([label, _safe(value)])
        summary.append(["Disclaimer", DISCLAIMER])
        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 90
        summary[summary.max_row][1].alignment = Alignment(wrap_text=True)

        measurements = workbook.create_sheet("Simplified Measurements")
        headers = [
            "Eye", "K1 flat (D)", "K2 steep (D)", "K2 status",
            "Pachymetry (µm)", "Pachymetry status", "Cylinder (D)",
            "Cylinder-magnitude status", "Image status", "Per-eye decision",
        ]
        measurements.append(headers)
        for cell in measurements[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="003366")
        for eye in self._eyes_by_laterality(screening_data).values():
            measurement = self._latest_measurements(eye)
            flags = self._flags_for_eye(eye, measurement)
            measurements.append([
                eye.get("laterality"), _safe(measurement.get("k1_d")), _safe(measurement.get("k2_d")), flags["k2"],
                _safe(measurement.get("pachymetry_um")), flags["pachymetry"], _safe(measurement.get("cylinder_d")),
                flags["cylinder"], _safe(eye.get("image_status") or eye.get("eye_result")), self._eye_decision(eye),
            ])
        for column in "ABCDEFGHIJ":
            measurements.column_dimensions[column].width = 23
        workbook.save(output)
        return str(output.resolve())

    # ------------------------------------------------------------------
    # Cumulative mass-screening register
    # ------------------------------------------------------------------

    REGISTER_HEADERS = (
        "Screening ID", "Screening date", "Age", "Sex", "Site", "Operator",
        "Outcome", "Action", "Referral priority", "Affected eye(s)",
        "OD image", "OD K1", "OD K2", "OD thickness", "OD cylinder", "OD decision",
        "OS image", "OS K1", "OS K2", "OS thickness", "OS cylinder", "OS decision",
        "Recorded at",
    )

    def _register_row(self, screening_data: dict[str, Any]) -> list[Any]:
        eyes = self._eyes_by_laterality(screening_data)
        row: list[Any] = [
            _safe(screening_data.get("screening_id")),
            _safe(screening_data.get("screening_date")),
            _safe(screening_data.get("age")),
            _safe(screening_data.get("sex")),
            _safe(screening_data.get("site")),
            _safe(screening_data.get("operator_id")),
            _safe(screening_data.get("overall_result")),
            _safe(screening_data.get("overall_action")),
            _safe(screening_data.get("referral_priority")),
            ", ".join(self._affected_eyes(screening_data)) or "—",
        ]
        for laterality in ("OD", "OS"):
            eye = eyes.get(laterality, {})
            measurement = self._latest_measurements(eye)
            row += [
                _safe(eye.get("image_status") or eye.get("eye_result")),
                _safe(measurement.get("k1_d")),
                _safe(measurement.get("k2_d")),
                _safe(measurement.get("pachymetry_um")),
                _safe(measurement.get("cylinder_d")),
                self._eye_decision(eye),
            ]
        row.append(datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
        return row

    def append_to_register(self, screening_data: dict[str, Any], register_path: str | Path) -> str:
        """Append one row per child to a cumulative mass-screening register.

        Re-exporting the same screening updates its existing row instead of
        adding a duplicate, so the register stays one row per child across a
        whole screening camp.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:  # pragma: no cover - dependency is declared
            raise RuntimeError("openpyxl is required for the local screening register.") from exc

        path = Path(register_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists():
            workbook = openpyxl.load_workbook(path)
            sheet = workbook["Register"] if "Register" in workbook.sheetnames else workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Register"
            sheet.append(list(self.REGISTER_HEADERS))
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="003366")
            for index in range(1, len(self.REGISTER_HEADERS) + 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 18

        row = self._register_row(screening_data)
        screening_id = row[0]
        target_row = None
        for existing in sheet.iter_rows(min_row=2):
            if existing and str(existing[0].value or "") == screening_id:
                target_row = existing[0].row
                break
        if target_row is None:
            sheet.append(row)
        else:
            for offset, value in enumerate(row):
                sheet.cell(row=target_row, column=offset + 1, value=value)
        workbook.save(path)
        return str(path.resolve())

    # ------------------------------------------------------------------
    # Detailed referral PDF
    # ------------------------------------------------------------------

    def _flags_for_eye(self, eye: dict[str, Any], measurement: dict[str, Any]) -> dict[str, str]:
        codes = set(eye.get("reason_codes") or [])
        return {
            "k2": "ABNORMAL" if "K2_ABOVE_46_8_D" in codes else ("MISSING" if measurement.get("k2_d") is None else "WITHIN THRESHOLD"),
            "pachymetry": "ABNORMAL" if "PACHYMETRY_BELOW_480_UM" in codes else ("MISSING" if measurement.get("pachymetry_um") is None else "WITHIN THRESHOLD"),
            "cylinder": "ABNORMAL" if "CYLINDER_MAGNITUDE_ABOVE_1_5_D" in codes else ("MISSING" if measurement.get("cylinder_d") is None else "WITHIN THRESHOLD"),
        }

    @staticmethod
    def _artifact_manifest(eye: dict[str, Any]) -> dict[str, dict[str, Any]]:
        manifest = eye.get("analysis_artifacts")
        if isinstance(manifest, dict):
            return manifest
        analyses = eye.get("image_analyses") or []
        if analyses and isinstance(analyses[-1].get("artifact_manifest"), dict):
            return analyses[-1]["artifact_manifest"]
        return {}

    def _verified_report_image(self, eye: dict[str, Any], filename: str) -> Path | None:
        """Return only an intact de-identified analysis artefact from this eye/run."""
        record = self._artifact_manifest(eye).get(filename)
        if not isinstance(record, dict):
            return None
        path_value = record.get("path")
        expected_hash = record.get("sha256")
        if not path_value or not expected_hash:
            return None
        if record.get("eye") not in (None, eye.get("laterality")):
            return None
        if record.get("source_image_hash") not in (None, "", eye.get("image_hash")):
            return None
        expected_run = eye.get("analysis_provenance_hash")
        if expected_run and record.get("provenance_hash") != expected_run:
            return None
        path = Path(str(path_value))
        if not path.is_file() or _sha256(path) != expected_hash:
            return None
        return path

    @staticmethod
    def _image_flowable(path: Path, label: str, styles, width_cm: float = 7.0):
        from reportlab.lib.units import cm
        from reportlab.platypus import Image as RLImage, KeepTogether, Paragraph, Spacer
        from PIL import Image as PILImage

        with PILImage.open(path) as image:
            width, height = image.size
        draw_width = width_cm * cm
        draw_height = draw_width * height / max(width, 1)
        max_height = 8.2 * cm
        if draw_height > max_height:
            draw_height = max_height
            draw_width = draw_height * width / max(height, 1)
        return KeepTogether([
            RLImage(str(path), width=draw_width, height=draw_height),
            Spacer(1, 0.08 * cm),
            Paragraph(_paragraph_text(label), styles["small"]),
            Spacer(1, 0.22 * cm),
        ])

    def _build_natural_reason_sentences(
        self, affected_eyes: list[str], eyes: dict[str, dict[str, Any]], protocol
    ) -> list[str]:
        """Build plain natural sentences from reason codes — no raw feature names exposed."""
        sentences: list[str] = []
        for laterality in affected_eyes:
            eye = eyes[laterality]
            codes = set(eye.get("reason_codes") or [])
            measurement = self._latest_measurements(eye)
            eye_label = "The right eye (OD)" if laterality == "OD" else "The left eye (OS)"

            if "K2_ABOVE_46_8_D" in codes:
                k2 = measurement.get("k2_d")
                if k2 is not None:
                    sentences.append(
                        f"{eye_label}: the K2 reading was {float(k2):.2f} D, which is above the "
                        f"configured screening threshold of {protocol.k2_abnormal_above_d:g} D."
                    )

            if "PACHYMETRY_BELOW_480_UM" in codes:
                pachy = measurement.get("pachymetry_um")
                if pachy is not None:
                    sentences.append(
                        f"{eye_label}: the pachymetry reading was {float(pachy):.0f} µm, which is below "
                        f"the configured screening threshold of {protocol.pachymetry_abnormal_below_um:g} µm."
                    )

            if "CYLINDER_MAGNITUDE_ABOVE_1_5_D" in codes:
                cyl = measurement.get("cylinder_d")
                if cyl is not None:
                    sentences.append(
                        f"{eye_label}: the cylinder magnitude was {abs(float(cyl)):.2f} D, which is above "
                        f"the configured screening threshold of {protocol.cylinder_magnitude_abnormal_above_d:g} D."
                    )

            if "IMAGE_CLASSIFIER_SUSPICIOUS" in codes:
                sentences.append(
                    f"{eye_label}: the KeraScan image showed a Placido ring pattern that was classified "
                    "as suspicious by a provisional, non-clinical geometry-consistency heuristic (not a "
                    "validated diagnostic classifier)."
                )

        if sentences:
            sentences.append(
                "This result is not a diagnosis. Screening results require interpretation by a qualified "
                "clinician in the context of the full clinical picture."
            )
        return sentences

    def _reason_rows(self, affected_eyes: list[str], eyes: dict[str, dict[str, Any]], protocol) -> list[list[str]]:
        rows: list[list[str]] = []
        thresholds = {
            "K2_ABOVE_46_8_D": f">{protocol.k2_abnormal_above_d:g} D",
            "PACHYMETRY_BELOW_480_UM": f"<{protocol.pachymetry_abnormal_below_um:g} µm",
            "CYLINDER_MAGNITUDE_ABOVE_1_5_D": f">{protocol.cylinder_magnitude_abnormal_above_d:g} D magnitude",
        }
        for laterality in affected_eyes:
            eye = eyes[laterality]
            codes = eye.get("reason_codes") or []
            measurement = self._latest_measurements(eye)
            if "IMAGE_CLASSIFIER_SUSPICIOUS" in codes:
                rows.append([
                    laterality, "KeraScan image", "Suspicious", "—",
                    "Image pattern classified as suspicious (provisional, non-clinical heuristic)",
                ])
            for code, (domain, key, reason) in MEASUREMENT_REASON_DETAILS.items():
                if code not in codes:
                    continue
                value = measurement.get(key)
                if code == "CYLINDER_MAGNITUDE_ABOVE_1_5_D" and value is not None:
                    observed = f"{abs(float(value)):.2f} D magnitude"
                elif code == "PACHYMETRY_BELOW_480_UM" and value is not None:
                    observed = f"{float(value):.0f} µm"
                else:
                    observed = f"{float(value):.2f} D" if value is not None else "—"
                rows.append([laterality, domain, observed, thresholds[code], reason])
        return rows

    def _report_identifier(self, screening_data: dict[str, Any], affected_eyes: list[str]) -> str:
        material = {
            "screening_id": screening_data.get("screening_id"),
            "affected_eyes": affected_eyes,
            "provenance": [
                (eye.get("laterality"), eye.get("analysis_provenance_hash"), eye.get("image_hash"))
                for eye in screening_data.get("eyes", []) if eye.get("laterality") in affected_eyes
            ],
        }
        return hashlib.sha256(json.dumps(material, sort_keys=True, default=str).encode("utf-8")).hexdigest()[:16].upper()

    @staticmethod
    def verify_pdf(path: str | Path) -> bool:
        """Basic local structural validation without adding a network dependency."""
        try:
            payload = Path(path).read_bytes()
        except OSError:
            return False
        return len(payload) > 1000 and payload.startswith(b"%PDF-") and b"%%EOF" in payload[-2048:]

    def generate_pdf(self, screening_data: dict[str, Any], output_path: str) -> str:
        """Generate the detailed referral PDF only for final child-level REFER."""
        if not self._is_refer_action(screening_data):
            raise ValueError("Detailed referral PDFs are generated only when the final child-level action is REFER.")
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                HRFlowable, KeepTogether, LongTable, PageBreak, Paragraph,
                SimpleDocTemplate, Spacer, Table, TableStyle,
            )
        except ImportError as exc:  # pragma: no cover - declared dependency
            raise RuntimeError("reportlab is required for local PDF generation.") from exc

        protocol = load_protocol()
        if tuple(protocol.detailed_pdf_for_actions) and "REFER" not in protocol.detailed_pdf_for_actions:
            raise ValueError("The configured protocol does not allow detailed PDF generation for REFER.")
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        eyes = self._eyes_by_laterality(screening_data)
        affected_eyes = self._affected_eyes(screening_data)
        if not affected_eyes:
            raise ValueError("A screen-positive referral PDF requires at least one affected eye.")
        if any(eye not in eyes for eye in affected_eyes):
            raise ValueError("Referral PDF eye provenance is incomplete.")

        report_id = self._report_identifier(screening_data, affected_eyes)
        generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        screening_timestamp = _safe(screening_data.get("screening_date"))
        if "T" not in screening_timestamp and screening_data.get("created_at"):
            screening_timestamp = f"{screening_timestamp} / recorded {screening_data['created_at']}"
        stylesheet = getSampleStyleSheet()
        styles = {
            "title": ParagraphStyle("KeraTitle", parent=stylesheet["Title"], fontName="Helvetica-Bold", fontSize=17, leading=21, textColor=colors.HexColor("#003366")),
            "subtitle": ParagraphStyle("KeraSub", parent=stylesheet["Normal"], fontSize=8, leading=11, textColor=colors.HexColor("#444444")),
            "section": ParagraphStyle("KeraSection", parent=stylesheet["Heading2"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#003366"), spaceBefore=8, spaceAfter=5),
            "normal": ParagraphStyle("KeraNormal", parent=stylesheet["Normal"], fontSize=8.4, leading=11),
            "small": ParagraphStyle("KeraSmall", parent=stylesheet["Normal"], fontSize=7.3, leading=9),
            "disclaimer": ParagraphStyle("KeraDisclaimer", parent=stylesheet["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#5B3D00")),
        }

        def cell(value: Any, style: str = "normal") -> Paragraph:
            return Paragraph(_paragraph_text(value), styles[style])

        def table_style(header: bool = True) -> TableStyle:
            commands = [
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9C7D4")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            if header:
                commands += [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            return TableStyle(commands)

        model_versions = []
        pipeline_versions = []
        for eye in eyes.values():
            if eye.get("model_version"):
                model_versions.append(str(eye["model_version"]))
            elif (eye.get("image_analyses") or [{}])[-1].get("model_version"):
                model_versions.append(str((eye.get("image_analyses") or [{}])[-1]["model_version"]))
            if eye.get("pipeline_version"):
                pipeline_versions.append(str(eye["pipeline_version"]))
        model_version = ", ".join(sorted(set(model_versions))) or "not available"
        pipeline_version = ", ".join(sorted(set(pipeline_versions))) or "not available"
        # Short hash prefixes: enough to tie the report to the stored analysis
        # record, without a wall of hex across the footer. The full hashes stay
        # in the JSON export and the local database.
        provenance_summary = "; ".join(
            f"{eye}: {_safe(eyes[eye].get('analysis_provenance_hash'))[:12]}" for eye in affected_eyes
        )

        def header_footer(canvas, doc) -> None:
            canvas.saveState()
            width, height = A4
            canvas.setFillColor(colors.HexColor("#003366"))
            canvas.rect(0, height - 1.1 * cm, width, 1.1 * cm, fill=True, stroke=False)
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica-Bold", 9)
            canvas.drawString(1.2 * cm, height - 0.72 * cm, "KeraScan School Corneal Screening Report")
            canvas.setFont("Helvetica", 7)
            canvas.drawRightString(width - 1.2 * cm, height - 0.72 * cm, f"Report ID {report_id}")
            canvas.setFillColor(colors.HexColor("#555555"))
            canvas.setFont("Helvetica", 7)
            canvas.drawString(1.2 * cm, 0.65 * cm, "Experimental initial screening aid — does not diagnose or exclude keratoconus.")
            canvas.drawRightString(width - 1.2 * cm, 0.65 * cm, f"Page {doc.page}")
            canvas.restoreState()

        story: list[Any] = [
            Paragraph("KeraScan School Corneal Screening Report", styles["title"]),
            Paragraph("Initial screening result — not a confirmed diagnosis", styles["subtitle"]),
            HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#003366"), spaceAfter=8),
            Paragraph("Screening information", styles["section"]),
        ]
        # Kept deliberately short: the identifiers a clinician needs to act on the
        # referral. Version/provenance strings are recorded in the JSON export and
        # summarised in one small line at the end rather than shown as a table.
        metadata = [
            [cell("Child/study ID", "small"), cell("Screening date", "small"), cell("Site", "small"), cell("Operator", "small")],
            [
                cell(screening_data.get("screening_id")),
                cell(screening_timestamp if screening_timestamp != "—" else generated_at),
                cell(screening_data.get("site")),
                cell(screening_data.get("operator_id")),
            ],
        ]
        meta_table = Table(metadata, colWidths=[4.2 * cm, 4.2 * cm, 4.2 * cm, 2.6 * cm], repeatRows=1)
        meta_table.setStyle(table_style())
        story += [meta_table, Spacer(1, 0.25 * cm), Paragraph("Screening outcome", styles["section"])]
        affected_label = "Both eyes (OD and OS)" if set(affected_eyes) == {"OD", "OS"} else (
            "Right eye (OD)" if affected_eyes[0] == "OD" else "Left eye (OS)"
        )
        outcome_wording = "Referral recommended — screening findings need specialist assessment"
        outcome = [
            [cell("Result", "small"), cell("Priority", "small"), cell("Affected eye(s)", "small")],
            [cell(outcome_wording), cell(screening_data.get("referral_priority")), cell(affected_label)],
        ]
        outcome_table = Table(outcome, colWidths=[7.5 * cm, 3.5 * cm, 4.2 * cm])
        outcome_table.setStyle(table_style())
        story += [outcome_table, Spacer(1, 0.10 * cm)]
        story.append(Paragraph(
            "Referral for corneal tomography and specialist assessment is advised.",
            ParagraphStyle("KeraReferral", parent=styles["normal"], textColor=colors.HexColor("#8B0000"), fontName="Helvetica-Bold"),
        ))
        story += [Spacer(1, 0.20 * cm), Paragraph("Why this child screened positive", styles["section"])]
        reasons = self._reason_rows(affected_eyes, eyes, protocol)
        if not reasons:
            raise ValueError("A referral report requires actual positive reason codes; none were found.")

        # Plain-English findings only. The observed-value/threshold detail these
        # sentences are built from is already in the measurements table below, so
        # the older reason-code table is not repeated here.
        natural_sentences = self._build_natural_reason_sentences(affected_eyes, eyes, protocol)
        for sentence in natural_sentences:
            story.append(Paragraph(_paragraph_text(f"• {sentence}"), styles["normal"]))
            story.append(Spacer(1, 0.08 * cm))

        story += [Spacer(1, 0.25 * cm), Paragraph("Measurements", styles["section"])]

        # Values with the out-of-threshold ones marked, in plain words. The
        # per-eye decision enum is deliberately not repeated here: the outcome
        # and the findings above already state it in readable form.
        ring_pattern_words = {
            "SUSPICIOUS": "irregular",
            "NORMAL_LIKE": "regular",
            "INDETERMINATE": "borderline",
        }
        measurement_data = [[cell(label, "small") for label in (
            "Eye", "K1", "K2", "Thickness", "Cylinder", "Ring pattern"
        )]]
        for laterality in ("OD", "OS"):
            eye = eyes.get(laterality, {})
            measurement = self._latest_measurements(eye)
            flags = self._flags_for_eye(eye, measurement)

            # Plain words rather than a warning glyph: the core PDF fonts have no
            # dingbat coverage, so a symbol here renders as a black box.
            def _value(key: str, fmt: str, unit: str, flag_key: str, direction: str) -> str:
                value = measurement.get(key)
                if value is None:
                    return "—"
                marker = f" ({direction})" if flags.get(flag_key) == "ABNORMAL" else ""
                return f"{format(float(value), fmt)} {unit}{marker}"

            cylinder = measurement.get("cylinder_d")
            cylinder_label = "—"
            if cylinder is not None:
                marker = " (high)" if flags.get("cylinder") == "ABNORMAL" else ""
                cylinder_label = f"{float(cylinder):.2f} D{marker}"
            status = eye.get("image_status") or eye.get("eye_result") or ""
            measurement_data.append([
                cell("Right" if laterality == "OD" else "Left"),
                cell(_value("k1_d", ".2f", "D", "k1", "high")),
                cell(_value("k2_d", ".2f", "D", "k2", "high")),
                cell(_value("pachymetry_um", ".0f", "µm", "pachymetry", "low")),
                cell(cylinder_label),
                cell(ring_pattern_words.get(str(status), "not available")),
            ])
        measurement_table = LongTable(measurement_data, colWidths=[1.6 * cm, 2.6 * cm, 2.6 * cm, 2.8 * cm, 2.6 * cm, 3.0 * cm], repeatRows=1)
        measurement_table.setStyle(table_style())
        story += [measurement_table, Spacer(1, 0.25 * cm)]

        story += [Paragraph("Affected eye image", styles["section"])]
        # One image per affected eye, in order of preference. The full artefact
        # set stays on disk and in the JSON export for anyone who needs it; the
        # referral itself carries only the single most useful picture.
        image_preference = (
            "clinician_comparison_panel.png",
            "observed_vs_concentric_reference.png",
            "tracked_rings_cartesian.png",
            "cropped_roi.png",
            "cropped_roi_centres.png",
            "directional_spacing.png",
        )
        image_labels = {
            "cropped_roi.png": "cropped corneal ring image",
            "cropped_roi_centres.png": "cropped ring image with centre",
            "tracked_rings_cartesian.png": "detected ring pattern",
            "directional_spacing.png": "ring-spacing pattern",
            "observed_vs_concentric_reference.png": "observed rings (solid) vs even-spacing reference (dashed)",
            "clinician_comparison_panel.png": "ring-pattern comparison",
        }
        report_image_count = 0
        chosen_by_eye: dict[str, tuple[str, Path]] = {}
        for laterality in affected_eyes:
            eye = eyes[laterality]
            chosen: tuple[str, Path] | None = None
            for filename in image_preference:
                artifact = self._verified_report_image(eye, filename)
                if artifact is not None:
                    chosen = (filename, artifact)
                    break
            if chosen is None:
                raise ValueError(
                    f"Referral PDF cannot be generated because the verified {laterality} "
                    "analysis image set is unavailable."
                )
            chosen_by_eye[laterality] = chosen
            filename, artifact = chosen
            eye_name = "Right eye (OD)" if laterality == "OD" else "Left eye (OS)"
            story.append(self._image_flowable(artifact, f"{eye_name} — {image_labels.get(filename, filename)}", styles))
            report_image_count += 1
        if report_image_count == 0:
            raise ValueError("Referral PDF cannot be generated because verified affected-eye analysis artefacts are unavailable.")

        # Appendix: every remaining verified analysis image for the affected
        # eye(s), so a reviewer working from the referral alone has the full
        # picture set for that child without going back to the local database.
        appendix: list[Any] = []
        for laterality in affected_eyes:
            eye = eyes[laterality]
            eye_name = "Right eye (OD)" if laterality == "OD" else "Left eye (OS)"
            shown = {name for name, _ in [chosen_by_eye[laterality]]} if laterality in chosen_by_eye else set()
            for filename in image_preference:
                if filename in shown:
                    continue
                artifact = self._verified_report_image(eye, filename)
                if artifact is None:
                    continue
                appendix.append(
                    self._image_flowable(artifact, f"{eye_name} — {image_labels.get(filename, filename)}", styles)
                )
        if appendix:
            story.append(PageBreak())
            story.append(Paragraph("Analysis images", styles["section"]))
            story.append(Paragraph(
                _paragraph_text(
                    "Image-space ring measurements retained for later review. These are "
                    "engineering comparisons, not corneal maps, and are not diagnostic."
                ),
                styles["small"],
            ))
            story.append(Spacer(1, 0.15 * cm))
            story += appendix

        story += [
            Spacer(1, 0.25 * cm),
            Paragraph(DISCLAIMER, styles["disclaimer"]),
            Spacer(1, 0.15 * cm),
            Paragraph(
                _paragraph_text(
                    f"Protocol {screening_data.get('protocol_version') or protocol.protocol_version} · "
                    f"analysis {pipeline_version} · method {model_version} · "
                    f"provenance {provenance_summary}"
                ),
                styles["small"],
            ),
        ]
        document = SimpleDocTemplate(
            str(output), pagesize=A4, topMargin=1.55 * cm, bottomMargin=1.3 * cm,
            leftMargin=1.15 * cm, rightMargin=1.15 * cm,
        )
        document.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
        if not self.verify_pdf(output):
            try:
                output.unlink()
            except OSError:
                pass
            raise RuntimeError("Generated local PDF did not pass structural validation.")
        return str(output.resolve())

    def generate_all_exports(self, screening_data: dict[str, Any], output_dir: str) -> dict[str, str | dict[str, str]]:
        """Export JSON/XLSX for any outcome; add detailed PDF only for REFER."""
        base = Path(output_dir)
        screening_id = str(screening_data.get("screening_id", "report"))
        results: dict[str, str | dict[str, str]] = {}
        errors: dict[str, str] = {}
        for name, method, suffix in (
            ("json", self.generate_json, "json"),
            ("excel", self.generate_excel, "xlsx"),
        ):
            try:
                results[name] = method(screening_data, str(base / f"{screening_id}.{suffix}"))
            except Exception as exc:  # pragma: no cover - retained diagnostics
                log.exception("%s export failed", name)
                errors[name] = str(exc)
        if self._is_refer_action(screening_data):
            try:
                results["pdf"] = self.generate_pdf(screening_data, str(base / f"{screening_id}.pdf"))
            except Exception as exc:
                log.exception("pdf export failed")
                errors["pdf"] = str(exc)
        if errors:
            results["errors"] = errors
        return results
