"""Safety and provenance tests for local screen-positive referral PDFs."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from PIL import Image, ImageDraw

from app.services.report_service import DISCLAIMER, ReportService, _sha256


@pytest.fixture
def svc():
    return ReportService()


def _make_assets(tmp_path: Path, eye: str, source_hash: str, provenance_hash: str) -> dict:
    manifest = {}
    # Include both the new preferred images and the old fallback images so
    # the test fixture works with both the new and legacy PDF generation paths.
    colors = {
        "cropped_roi.png": "#4477aa",
        "cropped_roi_centres.png": "#aa7744",
        "tracked_rings_cartesian.png": "#44aa77",
        "directional_spacing.png": "#aa4477",
        "observed_vs_concentric_reference.png": "#7744aa",
        "clinician_comparison_panel.png": "#44aacc",
    }
    for filename, color in colors.items():
        path = tmp_path / eye / filename
        path.parent.mkdir(parents=True, exist_ok=True)
        image = Image.new("RGB", (140, 90), color)
        ImageDraw.Draw(image).text((5, 5), f"{eye}-{filename}", fill="white")
        image.save(path)
        manifest[filename] = {
            "path": str(path), "sha256": _sha256(path), "eye": eye,
            "source_image_hash": source_hash, "provenance_hash": provenance_hash,
        }
    # This should never appear in the PDF.
    Image.new("RGB", (140, 90), "red").save(tmp_path / eye / "original_full_resolution.png")
    return manifest


def _eye(eye: str, *, image_status="SUSPICIOUS", decision="HIGH_RISK_SCREEN_POSITIVE", codes=None,
         assets=None, source_hash=None, provenance_hash=None):
    codes = codes or ["IMAGE_CLASSIFIER_SUSPICIOUS", "K2_ABOVE_46_8_D"]
    return {
        "laterality": eye,
        "eye_result": "SUSPICIOUS" if image_status == "SUSPICIOUS" else "NORMAL-LIKE",
        "image_status": image_status,
        "image_hash": source_hash,
        "analysis_provenance_hash": provenance_hash,
        "analysis_artifacts": assets or {},
        "quality_gradable": True,
        "quality_metrics": {"ring_tracking_confidence": 0.91},
        "geometry_validation_status": "PASS",
        "pipeline_version": "synthetic-pipeline-1",
        "model_version": "synthetic-model-1",
        "reason_codes": codes,
        "measurements": [{"k1_d": 43.0, "k2_d": 47.2, "pachymetry_um": 472, "cylinder_d": -2.0, "pachymetry_measurement_type": "device_reported"}],
        "decisions": [{"final_result": decision}],
        "image_analyses": [],
    }


@pytest.fixture
def positive_screening_data(tmp_path):
    od_hash = "a" * 64
    os_hash = "b" * 64
    od_run = "c" * 64
    os_run = "d" * 64
    od = _eye("OD", assets=_make_assets(tmp_path, "OD", od_hash, od_run), source_hash=od_hash, provenance_hash=od_run)
    os = _eye(
        "OS", image_status="NORMAL_LIKE", decision="SCREEN_NEGATIVE", codes=[],
        assets=_make_assets(tmp_path, "OS", os_hash, os_run), source_hash=os_hash, provenance_hash=os_run,
    )
    return {
        "screening_id": "SYNTHETIC-REF-001", "screening_date": "2026-08-21T09:30:00Z",
        "operator_id": "SYNTH-OP", "device_id": "SYNTH-DEVICE", "protocol_version": "kerascan-school-screening-provisional-1",
        "software_version": "phase4-school-screening", "overall_result": "SCREEN_POSITIVE", "overall_action": "REFER",
        "referral_priority": "PRIORITY_1", "affected_eyes": ["OD"], "eyes": [od, os],
    }


def test_referral_pdf_is_generated_and_structurally_valid(svc, positive_screening_data, tmp_path):
    path = svc.generate_pdf(positive_screening_data, str(tmp_path / "referral.pdf"))
    assert Path(path).exists()
    assert svc.verify_pdf(path)
    assert Path(path).stat().st_size > 3000


@pytest.mark.parametrize(
    ("decision", "action"),
    [
        ("SCREEN_NEGATIVE", "NO_IMMEDIATE_REFERRAL"),
        ("REPEAT_REQUIRED", "REPEAT_MEASUREMENT"),
        ("INCOMPLETE_SCREENING", "INCOMPLETE"),
    ],
)
def test_no_detailed_pdf_for_nonpositive_outcomes(svc, positive_screening_data, tmp_path, decision, action):
    data = dict(positive_screening_data, overall_result=decision, overall_action=action, affected_eyes=[])
    output = tmp_path / f"{decision}.pdf"
    with pytest.raises(ValueError, match="only"):
        svc.generate_pdf(data, str(output))
    assert not output.exists()


def test_reason_table_contains_actual_positive_criteria_only(svc, positive_screening_data):
    eyes = svc._eyes_by_laterality(positive_screening_data)
    rows = svc._reason_rows(["OD"], eyes, __import__("app.services.protocol", fromlist=["load_protocol"]).load_protocol())
    assert any(row[1] == "KeraScan image" for row in rows)
    assert any(row[1] == "K2" and row[2] == "47.20 D" for row in rows)
    assert not any(row[1] == "Pachymetry" for row in rows)
    assert not any(row[1] == "Cylinder" for row in rows)


def test_report_uses_only_affected_eye_analysis_images(svc, positive_screening_data, monkeypatch, tmp_path):
    calls = []
    original = svc._verified_report_image

    def capture(eye, filename):
        calls.append((eye["laterality"], filename))
        return original(eye, filename)

    monkeypatch.setattr(svc, "_verified_report_image", capture)
    svc.generate_pdf(positive_screening_data, str(tmp_path / "affected.pdf"))
    # Must only attempt images for the affected eye (OD), never OS
    assert {eye for eye, _ in calls} == {"OD"}
    # The simplified report carries a single image per affected eye, so lookup
    # stops at the first artefact that verifies rather than collecting a set.
    assert len(calls) >= 1


def test_report_refuses_swapped_or_hash_mismatched_images(svc, positive_screening_data, tmp_path):
    # Swap all images so none can be verified for OD — should raise
    swapped = json.loads(json.dumps(positive_screening_data))
    for filename in list(swapped["eyes"][0]["analysis_artifacts"]):
        swapped["eyes"][0]["analysis_artifacts"][filename]["eye"] = "OS"
    with pytest.raises(ValueError, match="verified OD|unavailable"):
        svc.generate_pdf(swapped, str(tmp_path / "swapped.pdf"))

    # Hash mismatch on all images — should raise
    mismatched = json.loads(json.dumps(positive_screening_data))
    for filename in list(mismatched["eyes"][0]["analysis_artifacts"]):
        mismatched["eyes"][0]["analysis_artifacts"][filename]["sha256"] = "0" * 64
    with pytest.raises(ValueError, match="verified OD|unavailable"):
        svc.generate_pdf(mismatched, str(tmp_path / "mismatched.pdf"))


def test_pdf_excludes_full_face_source_and_local_paths(svc, positive_screening_data, tmp_path):
    path = svc.generate_pdf(positive_screening_data, str(tmp_path / "privacy.pdf"))
    payload = Path(path).read_bytes()
    assert b"original_full_resolution" not in payload
    assert str(tmp_path).encode() not in payload


def test_pdf_requires_verified_analysis_images(svc, positive_screening_data, tmp_path):
    """PDF must require at least one verified analysis image per affected eye."""
    # Remove all analysis artifacts for OD — should raise
    positive_screening_data["eyes"][0]["analysis_artifacts"] = {}
    with pytest.raises(ValueError, match="verified OD|unavailable"):
        svc.generate_pdf(positive_screening_data, str(tmp_path / "empty_artifacts.pdf"))


def test_json_is_path_redacted_and_disclaimer_is_safe(svc, positive_screening_data, tmp_path):
    path = svc.generate_json(positive_screening_data, str(tmp_path / "record.json"))
    payload = Path(path).read_text()
    assert "does not diagnose or exclude keratoconus" in payload
    assert str(tmp_path) not in payload
    assert DISCLAIMER in payload


def test_excel_contains_only_simplified_measurement_sheet(svc, positive_screening_data, tmp_path):
    import openpyxl

    path = svc.generate_excel(positive_screening_data, str(tmp_path / "record.xlsx"))
    workbook = openpyxl.load_workbook(path)
    assert "Summary" in workbook.sheetnames
    assert "Simplified Measurements" in workbook.sheetnames
    assert "Measurements" not in workbook.sheetnames


def test_generate_all_exports_skips_pdf_when_not_referred(svc, positive_screening_data, tmp_path):
    data = dict(positive_screening_data, overall_result="SCREEN_NEGATIVE", overall_action="NO_IMMEDIATE_REFERRAL", affected_eyes=[])
    exports = svc.generate_all_exports(data, str(tmp_path))
    assert "pdf" not in exports
    assert Path(exports["json"]).exists()
    assert Path(exports["excel"]).exists()


def test_pdf_wording_is_screening_not_diagnosis(svc, positive_screening_data):
    source = Path(__import__("app.services.report_service", fromlist=["__file__"]).__file__).read_text().lower()
    assert "keratoconus confirmed" not in source
    assert "keratoconus diagnosed" not in source
    assert "defected patient" not in source
    # Report now uses correct screening terminology
    assert "school corneal screening report" in source
    # Must contain correct screening outcome language (not raw diagnostic label)
    assert "screen-positive" in source or "screen positive" in source


# ---------------------------------------------------------------------------
# Cumulative mass-screening register
# ---------------------------------------------------------------------------

def test_register_accumulates_one_row_per_child(svc, positive_screening_data, tmp_path):
    import openpyxl

    register = tmp_path / "screening_register.xlsx"
    first = json.loads(json.dumps(positive_screening_data))
    first["screening_id"] = "CAMP-001"
    second = json.loads(json.dumps(positive_screening_data))
    second["screening_id"] = "CAMP-002"

    svc.append_to_register(first, register)
    svc.append_to_register(second, register)

    sheet = openpyxl.load_workbook(register)["Register"]
    ids = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert ids == ["CAMP-001", "CAMP-002"]
    assert sheet.cell(row=1, column=1).value == "Screening ID"


def test_register_updates_rather_than_duplicates_the_same_child(svc, positive_screening_data, tmp_path):
    import openpyxl

    register = tmp_path / "screening_register.xlsx"
    svc.append_to_register(positive_screening_data, register)
    svc.append_to_register(positive_screening_data, register)

    sheet = openpyxl.load_workbook(register)["Register"]
    ids = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert len(ids) == 1, "re-exporting the same child must not add a duplicate row"
