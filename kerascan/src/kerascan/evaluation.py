"""Local, patient-grouped evaluation. Never uploads or copies source images."""
from __future__ import annotations

import json
import os
import tempfile
import math
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

os.environ.setdefault("MPLCONFIGDIR", str(Path(tempfile.gettempdir()) / "kerascan-matplotlib"))
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors
from sklearn.calibration import calibration_curve
from sklearn.metrics import (accuracy_score, average_precision_score, brier_score_loss, confusion_matrix,
                             precision_recall_curve, precision_recall_fscore_support, roc_auc_score, roc_curve,
                             f1_score, balanced_accuracy_score)

from .config import EngineConfig, ModelConfig
from .features import FEATURE_ORDER, feature_vector
from .inference import EXPERIMENTAL, KerascanEngine
from .manifest import ManifestRecord
from .model_bundle import ModelBundle
from .privacy import ensure_local_output, sha256_text

EVALUATION_VERSION = "phase3-0.1.0"
REFERENCE_TO_ENGINE = {"NORMAL": "NORMAL-LIKE", "SUSPICIOUS": "SUSPICIOUS", "UNGRADABLE": "UNGRADABLE"}


@dataclass
class EvaluationCase:
    patient_id: str
    eye: str
    session_id: str
    reference_label: str
    prediction: str
    score: float | None
    quality_flags: list[str]
    failure_category: str
    source: str = "image"
    k1_d: float | None = None
    k2_d: float | None = None
    pachymetry_um: float | None = None
    cylinder_d: float | None = None
    roi_success: bool = False
    outer_ring_radius_fraction: float | None = None
    angular_coverage: float | None = None
    missing_ring_fraction: float | None = None
    tracking_confidence: float | None = None

    def public_dict(self) -> dict:
        return asdict(self)


def _float_or_none(value: str | None) -> float | None:
    try: return float(value) if value not in (None, "") else None
    except (TypeError, ValueError): return None


def categorize_failure(reference: str, prediction: str, flags: list[str], tracking: dict | None) -> str:
    flags = set(flags or [])
    if "non_ring_or_incorrect_input" in flags: return "Non-ring input"
    if "blur" in flags: return "Blur"
    if "glare_or_saturation" in flags: return "Glare"
    if "possible_eyelid_or_eyelash_obstruction" in flags: return "Occlusion"
    if "low_contrast" in flags: return "Low contrast"
    if "sensor_noise" in flags: return "Excessive noise"
    if "placido_pattern_too_small" in flags or "pattern_off_centre" in flags: return "ROI detection failure"
    if "excessive_missing_or_broken_rings" in flags: return "Missing rings"
    if tracking and tracking.get("identity_shift_fraction", 0) > .25: return "Ring-identity shift"
    if prediction == "UNGRADABLE": return "Segmentation failure"
    if reference == "SUSPICIOUS" and prediction == "NORMAL-LIKE": return "Image-model false negative"
    if reference == "NORMAL" and prediction == "SUSPICIOUS": return "Image-model false positive"
    return "Unknown"


def infer_cases(records: list[ManifestRecord], bundle: ModelBundle) -> list[EvaluationCase]:
    """Run local inference. Calls receive no output directory, so no source image is copied."""
    config = EngineConfig(model=ModelConfig(random_seed=20260821, decision_threshold=float(bundle.threshold)))
    engine = KerascanEngine(config)
    engine.model = bundle.estimator
    engine.model_metadata = {**bundle.metadata(), "model_hash": bundle.model_hash, "feature_schema_hash": bundle.feature_schema_hash,
                             "score_status": EXPERIMENTAL}
    results: list[EvaluationCase] = []
    for record in records:
        if record.reference_label == "EXCLUDE": continue
        try:
            result = engine.analyze(record.image_path)  # No output folder: no image copies/previews.
            prediction = result["screening_result"]
            score = result.get("prototype_score")
            flags = list((result.get("quality") or {}).get("flags") or [])
            tracking = result.get("tracking")
            roi=result.get("roi") or {}; features=result.get("features") or {}; tracking=tracking or {}
            source_fraction=(result.get("quality") or {}).get("metrics",{}).get("source_pattern_radius_fraction")
            results.append(EvaluationCase(record.patient_id, record.eye, record.session_id, record.reference_label,
                prediction, float(score) if score is not None else None, flags,
                categorize_failure(record.reference_label, prediction, flags, tracking),
                k1_d=_float_or_none(record.extras.get("k1_d")), k2_d=_float_or_none(record.extras.get("k2_d")),
                pachymetry_um=_float_or_none(record.extras.get("pachymetry_um")),
                cylinder_d=_float_or_none(record.extras.get("cylinder_d")), roi_success=bool(roi.get("confidence",0) > 0),
                outer_ring_radius_fraction=float(source_fraction) if source_fraction is not None else None,
                angular_coverage=features.get("angular_coverage"), missing_ring_fraction=features.get("missing_ring_fraction"),
                tracking_confidence=tracking.get("confidence")))
        except Exception:
            # Deliberately avoid emitting private paths/exceptions in aggregate output.
            results.append(EvaluationCase(record.patient_id, record.eye, record.session_id, record.reference_label,
                "UNGRADABLE", None, ["analysis_error"], "ROI detection failure"))
    return results


def _binary_arrays(cases: list[EvaluationCase]) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    valid = [case for case in cases if case.reference_label in {"NORMAL", "SUSPICIOUS"} and case.prediction in {"NORMAL-LIKE", "SUSPICIOUS"}]
    truth = np.array([case.reference_label == "SUSPICIOUS" for case in valid], dtype=int)
    pred = np.array([case.prediction == "SUSPICIOUS" for case in valid], dtype=int)
    scores = np.array([case.score if case.score is not None else np.nan for case in valid], dtype=float)
    return truth, pred, scores


def metric_summary(cases: list[EvaluationCase]) -> dict[str, Any]:
    included = [case for case in cases if case.reference_label != "EXCLUDE"]
    truth, pred, scores = _binary_arrays(included)
    all_correct = [REFERENCE_TO_ENGINE.get(case.reference_label) == case.prediction for case in included]
    cm = confusion_matrix(truth, pred, labels=[0, 1]).tolist() if len(truth) else [[0, 0], [0, 0]]
    tn, fp, fn, tp = np.asarray(cm).ravel()
    sensitivity = tp / (tp + fn) if tp + fn else None
    specificity = tn / (tn + fp) if tn + fp else None
    ppv = tp / (tp + fp) if tp + fp else None
    npv = tn / (tn + fn) if tn + fn else None
    per_precision, per_recall, per_f1, per_support = precision_recall_fscore_support(truth, pred, labels=[0,1], zero_division=0) if len(truth) else (np.zeros(2),)*4
    has_both_classes = len(np.unique(truth)) == 2
    finite_scores = bool(len(scores) and np.all(np.isfinite(scores)))
    values=lambda name: [float(getattr(case,name)) for case in included if getattr(case,name) is not None]
    return {
        "total_cases": len(included), "total_patients": len({case.patient_id for case in included}),
        "per_eye_counts": dict(Counter(case.eye for case in included)), "per_child_counts": {"patients": len({case.patient_id for case in included})},
        "class_distribution": dict(Counter(case.reference_label for case in included)),
        "accuracy": float(np.mean(all_correct)) if all_correct else None,
        "gradable_classification_cases": int(len(truth)), "balanced_accuracy": float(balanced_accuracy_score(truth,pred)) if has_both_classes else None,
        "sensitivity": sensitivity, "specificity": specificity, "ppv": ppv, "npv": npv,
        "macro_f1": float(f1_score(truth,pred,average="macro",zero_division=0)) if len(truth) else None,
        "per_class": {"NORMAL": {"precision": float(per_precision[0]), "recall": float(per_recall[0]), "f1": float(per_f1[0]), "support": int(per_support[0])},
                      "SUSPICIOUS": {"precision": float(per_precision[1]), "recall": float(per_recall[1]), "f1": float(per_f1[1]), "support": int(per_support[1])}},
        "confusion_matrix": cm,
        "roc_auc": float(roc_auc_score(truth,scores)) if has_both_classes and finite_scores else None,
        "pr_auc": float(average_precision_score(truth,scores)) if has_both_classes and finite_scores else None,
        "brier_score": float(brier_score_loss(truth,scores)) if has_both_classes and finite_scores else None,
        "ungradable_rate": float(np.mean([case.prediction == "UNGRADABLE" for case in included])) if included else None,
        "referral_rate": float(np.mean([case.prediction in {"SUSPICIOUS", "UNGRADABLE"} for case in included])) if included else None,
        "roi_detection_success_rate": float(np.mean([case.roi_success for case in included])) if included else None,
        "mean_outer_ring_radius_fraction": float(np.mean(values("outer_ring_radius_fraction"))) if values("outer_ring_radius_fraction") else None,
        "mean_angular_coverage": float(np.mean(values("angular_coverage"))) if values("angular_coverage") else None,
        "mean_missing_ring_fraction": float(np.mean(values("missing_ring_fraction"))) if values("missing_ring_fraction") else None,
        "mean_tracking_confidence": float(np.mean(values("tracking_confidence"))) if values("tracking_confidence") else None,
    }


def grouped_bootstrap_ci(cases: list[EvaluationCase], *, iterations: int = 1000, seed: int = 20260821) -> dict[str, dict[str, float | None]]:
    patients = sorted({case.patient_id for case in cases})
    if len(patients) < 2: return {key: {"low": None, "high": None} for key in ("accuracy","sensitivity","specificity","balanced_accuracy","macro_f1","npv")}
    by_patient = {patient: [case for case in cases if case.patient_id == patient] for patient in patients}
    rng = np.random.default_rng(seed); collected: dict[str,list[float]] = {key: [] for key in ("accuracy","sensitivity","specificity","balanced_accuracy","macro_f1","npv")}
    for _ in range(iterations):
        sample = [case for patient in rng.choice(patients, size=len(patients), replace=True) for case in by_patient[patient]]
        metrics = metric_summary(sample)
        for key in collected:
            if metrics.get(key) is not None: collected[key].append(float(metrics[key]))
    return {key: ({"low": float(np.percentile(values,2.5)), "high": float(np.percentile(values,97.5))} if values else {"low":None,"high":None}) for key,values in collected.items()}


def target_assessment(metrics: dict[str, Any]) -> dict[str, Any]:
    """Describe evidence without inventing a performance or deployment target.

    Clinical acceptance criteria belong in an approved validation protocol, not
    in application code.  Even strong locked-test metrics do not automatically
    establish suitability for school screening.
    """
    return {
        "target_accuracy": None,
        "observed_locked_test_accuracy": metrics.get("accuracy"),
        "target_achieved": "NOT_ASSESSED",
        "screening_suitable_by_target": "NOT_DETERMINED",
        "reason": (
            "No hard-coded performance target is applied. Clinical suitability requires "
            "a prospectively approved validation protocol, confidence intervals, failure-rate "
            "assessment, and independent clinical review."
        ),
    }


def screening_system_comparison(
    cases: list[EvaluationCase], *, referral_engine: Any | None = None
) -> dict[str, Any]:
    """Compare image-only, quantitative, and combined screening behaviour.

    The versioned application referral engine is injected by the local evaluation
    CLI.  If it is unavailable, or the manifest lacks a complete canonical
    measurement set, the comparison is omitted rather than guessed from fallback
    constants.  This function never trains or modifies the image classifier.
    """
    image = metric_summary(cases)
    if referral_engine is None:
        unavailable = {
            "status": "not_evaluable",
            "reason": "The versioned screening protocol and referral engine were not supplied.",
        }
        return {
            "image_model": image,
            "measurement_rules": unavailable,
            "combined_system": unavailable,
        }

    measured = [
        case
        for case in cases
        if None not in (case.k1_d, case.k2_d, case.pachymetry_um, case.cylinder_d)
    ]
    if not measured:
        unavailable = {
            "status": "not_evaluable",
            "reason": (
                "Manifest contains no complete canonical K1, K2, pachymetry, and "
                "cylinder measurement set."
            ),
        }
        return {
            "image_model": image,
            "measurement_rules": unavailable,
            "combined_system": unavailable,
            "protocol_version": referral_engine.get_protocol_version(),
            "thresholds": referral_engine.thresholds,
        }

    rules: list[EvaluationCase] = []
    combined: list[EvaluationCase] = []
    combined_outcomes: Counter[str] = Counter()
    measurement_outcomes: Counter[str] = Counter()

    def prediction_for_result(result: Any) -> str:
        if result.action == "REFER":
            return "SUSPICIOUS"
        if result.decision == "SCREEN_NEGATIVE":
            return "NORMAL-LIKE"
        return "UNGRADABLE"

    for case in measured:
        measurements = {
            "k1_d": case.k1_d,
            "k2_d": case.k2_d,
            "pachymetry_um": case.pachymetry_um,
            "cylinder_d": case.cylinder_d,
        }
        measurement_result = referral_engine.evaluate_eye(
            case.eye,
            "NORMAL-LIKE",
            measurements,
            image_status="NORMAL_LIKE",
        )
        measurement_outcomes[measurement_result.decision] += 1
        if measurement_result.flags is None or measurement_result.action == "INCOMPLETE":
            measurement_prediction = "UNGRADABLE"
        elif measurement_result.flags.abnormal_measurement_count:
            # Measurement-only comparison records any abnormal domain as a
            # positive quantitative signal; the combined matrix still retains
            # the distinct REPEAT_REQUIRED state for one isolated abnormality.
            measurement_prediction = "SUSPICIOUS"
        else:
            measurement_prediction = "NORMAL-LIKE"
        rules.append(
            EvaluationCase(
                case.patient_id,
                case.eye,
                case.session_id,
                case.reference_label,
                measurement_prediction,
                None,
                [],
                "Unknown",
                source="configured_measurement_rules",
            )
        )

        image_status = {
            "NORMAL-LIKE": "NORMAL_LIKE",
            "SUSPICIOUS": "SUSPICIOUS",
            "ANALYSIS_BLOCKED": "ANALYSIS_BLOCKED",
        }.get(case.prediction, "IMAGE_REJECTED")
        combined_result = referral_engine.evaluate_eye(
            case.eye,
            case.prediction,
            measurements,
            image_status=image_status,
        )
        combined_outcomes[combined_result.decision] += 1
        combined.append(
            EvaluationCase(
                case.patient_id,
                case.eye,
                case.session_id,
                case.reference_label,
                prediction_for_result(combined_result),
                case.score,
                case.quality_flags,
                case.failure_category,
                source="configured_combined_screening_matrix",
            )
        )

    return {
        "image_model": image,
        "measurement_rules": metric_summary(rules),
        "combined_system": metric_summary(combined),
        "measurement_outcomes": dict(measurement_outcomes),
        "combined_outcomes": dict(combined_outcomes),
        "protocol_version": referral_engine.get_protocol_version(),
        "thresholds": referral_engine.thresholds,
        "comparison_scope": (
            "Per-eye configured screening outcomes; child-level escalation is applied "
            "by the same versioned application referral engine."
        ),
    }


def _safe_sheet(workbook: Workbook, name: str):
    return workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet(name)


def write_evaluation_outputs(output_dir: str | Path, *, cases: list[EvaluationCase], metrics: dict[str, Any],
                             confidence_intervals: dict[str, Any], provenance: dict[str, Any],
                             referral_engine: Any | None = None) -> dict[str, str]:
    """Write local aggregate outputs only. No raw images or source paths are included."""
    output = ensure_local_output(output_dir)
    assessment = target_assessment(metrics)
    comparison = screening_system_comparison(cases, referral_engine=referral_engine)
    aggregate = {"evaluation_version": EVALUATION_VERSION, "provenance": provenance, "metrics": metrics,
                 "confidence_intervals_95": confidence_intervals, "target_assessment": assessment,
                 "screening_system_comparison": comparison,
                 "warning": "Do not modify or retune the model based on locked-test results. Any subsequent model must be evaluated on a new untouched test set."}
    aggregate_path = output / "aggregate_metrics.json"; aggregate_path.write_text(json.dumps(aggregate, indent=2, sort_keys=True), encoding="utf-8")
    hash_path = output / "evaluation_manifest_hash.json"; hash_path.write_text(json.dumps(provenance,indent=2,sort_keys=True),encoding="utf-8")

    workbook = Workbook(); summary = workbook.active; summary.title = "Aggregate Metrics"
    summary.append(["Metric", "Value"])
    for key, value in metrics.items():
        if key not in {"per_class", "confusion_matrix", "class_distribution", "per_eye_counts", "per_child_counts"}:
            summary.append([key, value if not isinstance(value, (dict, list)) else json.dumps(value)])
    summary.append(["predefined_performance_target", assessment["target_accuracy"] or "not configured"])
    summary.append(["automated_suitability_assessment", assessment["screening_suitable_by_target"]])
    for key, interval in confidence_intervals.items(): summary.append([f"{key}_95_ci", json.dumps(interval)])
    cm_sheet = workbook.create_sheet("Confusion Matrix"); cm_sheet.append(["Reference / Prediction", "NORMAL-LIKE", "SUSPICIOUS"])
    cm_sheet.append(["NORMAL", *metrics["confusion_matrix"][0]]); cm_sheet.append(["SUSPICIOUS", *metrics["confusion_matrix"][1]])
    wb_path = output / "aggregate_metrics.xlsx"; workbook.save(wb_path)
    subgroup = Workbook(); subgroup.active.title = "Subgroups"; subgroup.active.append(["Subgroup", "Cases", "Patients", "Accuracy", "Ungradable rate"])
    for column in ("eye",):
        for value in sorted({getattr(case,column) for case in cases}):
            group = [case for case in cases if getattr(case,column) == value]; group_metrics = metric_summary(group)
            subgroup.active.append([f"{column}={value}", len(group), len({case.patient_id for case in group}), group_metrics["accuracy"], group_metrics["ungradable_rate"]])
    subgroup_path=output/"subgroup_metrics.xlsx"; subgroup.save(subgroup_path)
    failures = Workbook(); failures.active.title="Failure Summary"; failures.active.append(["Failure category","Count","Rate"])
    counts=Counter(case.failure_category for case in cases if case.failure_category != "Unknown")
    for category,count in sorted(counts.items()): failures.active.append([category,count,count/max(len(cases),1)])
    failure_path=output/"failure_summary.xlsx"; failures.save(failure_path)

    _write_plots(output, cases, metrics)
    _write_pdf(output / "final_screening_evaluation.pdf", metrics, confidence_intervals, provenance, assessment, counts, comparison)
    return {"aggregate_metrics_json":str(aggregate_path),"aggregate_metrics_xlsx":str(wb_path),"confusion_matrix":str(output/"confusion_matrix.png"),
            "roc_curve":str(output/"roc_curve.png"),"precision_recall_curve":str(output/"precision_recall_curve.png"),"calibration_curve":str(output/"calibration_curve.png"),
            "subgroup_metrics":str(subgroup_path),"failure_summary":str(failure_path),"evaluation_manifest_hash":str(hash_path),"final_pdf":str(output/"final_screening_evaluation.pdf")}


def _placeholder_plot(path: Path, title: str, text: str) -> None:
    fig,ax=plt.subplots(figsize=(6,4)); ax.axis("off"); ax.set_title(title); ax.text(.5,.5,text,ha="center",va="center",wrap=True); fig.tight_layout(); fig.savefig(path,dpi=150); plt.close(fig)


def _write_plots(output: Path, cases: list[EvaluationCase], metrics: dict[str,Any]) -> None:
    cm=np.asarray(metrics["confusion_matrix"]); fig,ax=plt.subplots(figsize=(5,4)); image=ax.imshow(cm,cmap="Blues")
    for i in range(2):
        for j in range(2): ax.text(j,i,str(cm[i,j]),ha="center",va="center")
    ax.set(xticks=[0,1],yticks=[0,1],xticklabels=["NORMAL-LIKE","SUSPICIOUS"],yticklabels=["NORMAL","SUSPICIOUS"],xlabel="Prediction",ylabel="Reference",title="Confusion matrix")
    fig.colorbar(image,ax=ax); fig.tight_layout(); fig.savefig(output/"confusion_matrix.png",dpi=150); plt.close(fig)
    truth,pred,scores=_binary_arrays(cases)
    if len(np.unique(truth)) == 2 and len(scores) and np.all(np.isfinite(scores)):
        fpr,tpr,_=roc_curve(truth,scores); fig,ax=plt.subplots(figsize=(5,4)); ax.plot(fpr,tpr,label=f"AUC={metrics['roc_auc']:.3f}"); ax.plot([0,1],[0,1],"--",color="grey"); ax.legend(); ax.set(xlabel="False positive rate",ylabel="True positive rate",title="ROC curve"); fig.tight_layout();fig.savefig(output/"roc_curve.png",dpi=150);plt.close(fig)
        precision,recall,_=precision_recall_curve(truth,scores);fig,ax=plt.subplots(figsize=(5,4));ax.plot(recall,precision,label=f"AP={metrics['pr_auc']:.3f}");ax.legend();ax.set(xlabel="Recall",ylabel="Precision",title="Precision–recall curve");fig.tight_layout();fig.savefig(output/"precision_recall_curve.png",dpi=150);plt.close(fig)
        observed,predicted=calibration_curve(truth,scores,n_bins=min(10,len(scores)),strategy="quantile");fig,ax=plt.subplots(figsize=(5,4));ax.plot(predicted,observed,"o-");ax.plot([0,1],[0,1],"--",color="grey");ax.set(xlabel="Mean predicted probability",ylabel="Observed suspicious rate",title="Calibration curve");fig.tight_layout();fig.savefig(output/"calibration_curve.png",dpi=150);plt.close(fig)
    else:
        for name,title in (("roc_curve.png","ROC curve"),("precision_recall_curve.png","Precision–recall curve"),("calibration_curve.png","Calibration curve")):
            _placeholder_plot(output/name,title,"Unavailable: the evaluated gradable cases do not contain both reference classes with probability scores.")


def _write_pdf(path: Path, metrics: dict[str,Any], intervals: dict[str,Any], provenance: dict[str,Any], assessment: dict[str,Any], failures: Counter, comparison: dict[str,Any]) -> None:
    styles=getSampleStyleSheet(); story=[Paragraph("KERASCAN Phase 3 Screening Evaluation",styles["Title"]),Spacer(1,12)]
    story += [Paragraph("Project objective",styles["Heading2"]),Paragraph("Portable initial school screening. This local report contains aggregate, de-identified evaluation results only; it contains no raw images or source paths.",styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Data partitions and leakage controls",styles["Heading2"]),Paragraph(f"Manifest hash: {provenance.get('test_manifest_hash','unknown')}<br/>Patient-level split checks: {provenance.get('patient_leakage_check','unknown')}<br/>Execution: {provenance.get('executed_at','unknown')}",styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Method attribution",styles["Heading2"]),Paragraph("Traditional contrast/edge mire segmentation, radial localisation, and spatial tracking are independently implemented with limited SmartKC-inspired methodology. See the local SmartKC MIT attribution and modification record.",styles["BodyText"]),Spacer(1,8)]
    table=[["Metric","Value"]]+[[key,str(metrics.get(key))] for key in ("total_cases","total_patients","accuracy","sensitivity","specificity","balanced_accuracy","macro_f1","roc_auc","pr_auc","brier_score","ungradable_rate","referral_rate")]
    t=Table(table,colWidths=[220,220]);t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#003366")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.25,colors.grey)]));story += [Paragraph("Classification and quality metrics",styles["Heading2"]),t,Spacer(1,8)]
    story += [Paragraph("Confidence intervals",styles["Heading2"]),Paragraph(json.dumps(intervals,sort_keys=True),styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Failure analysis",styles["Heading2"]),Paragraph(json.dumps(dict(failures),sort_keys=True),styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Screening-system comparison",styles["Heading2"]),Paragraph(json.dumps(comparison,sort_keys=True),styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Evidence assessment",styles["Heading2"]),Paragraph(
        f"Predefined performance target: not configured<br/>"
        f"Observed locked-test accuracy: {assessment.get('observed_locked_test_accuracy')}<br/>"
        f"Automated screening-suitability decision: {assessment.get('screening_suitable_by_target')}<br/>"
        f"{assessment.get('reason')}", styles["BodyText"]),Spacer(1,8)]
    story += [Paragraph("Field-readiness conclusion",styles["Heading2"]),Paragraph("KERASCAN is intended as a portable initial screening aid. The system identifies children who may require further corneal evaluation. It does not diagnose keratoconus and does not replace tomography or assessment by a qualified clinician.",styles["BodyText"])]
    SimpleDocTemplate(str(path),pagesize=A4,rightMargin=36,leftMargin=36,topMargin=36,bottomMargin=36).build(story)
